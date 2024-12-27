import openpyxl
import tkinter as tk
from tkinter import filedialog

root = tk.Tk()
root.withdraw()

path = filedialog.askopenfilename(
    title="Select an Excel File",
    filetypes=[("Excel Files", "*.xls *.xlsx")]
)

menu = {
    "Cheesy Bread" : f'=(C{cell.row}*4+D2*7)/16',
    "Mozz Cheese Bread" : f'=(C{cell.row}*4+D2*7)/16',
    "Pepperoni & Cheese" : f'=(C{cell.row}*4+D2*7)/16',
    "Ultimate 3 Cheese" : f'=(C{cell.row}*4+D2*7)/16',
    "Calzone" : f'=C{cell.row}*3/16',
    "Paulzone" : f'=C{cell.row}*3/16',
    "Catering Wings" : '', # No Equation for this yet
    "Wings" : '', # No Equation for this yet
    "Dessert" : '', # No Equation for this yet
    "Big Double Chocolate Cake" : f'C{cell.row}',
    "Giant Cookie" : f'C{cell.row}',
    "Oreo Cheesecake" : f'C{cell.row}',
    "Red Velvet Cake" : f'C{cell.row}',
    "Strawberry Cheesecake" : f'C{cell.row}',
    "Jumbo Wings (8pc)" : '', # No Equation for this yet
    "Drinks" : '', # Equation is self evident
    "Drinks Totals" : '', # Equation is self evident
    "Pizza" : '', 
    "Pizza Totals" : f'=(C{cell.row}*2.5+D{cell.row}*4+E{cell.row}*5.5+F{cell.row}*7.5+G{cell.row}*22+H{cell.row}*4)/16',
    "Slice" : '',
    "2 Slices" : f'=C{cell.row}*5/16',
    "Pizza Slice" : f'=C{cell.row}*2.5/16',
    "Subs" : '',
    "Ham" : '',
    "Italian" : '',
    "Italian Bake Sandwich" : '',
    "Turkey Club" : '',
    "Jumbo Wings (16)" : '', 
    "Jumbo Wings (6)" : '',
    "Jumbo Wings (8)" : '',
    "Bread" : '',
    "Bacon" : '',
    "Green Pepper" : '', 
    "Half Bacon" : '',
    "Hamburger" : '',
    "Burger" : '',
    "Onions" : '',
    "Cheese" : '',
    "Ground Beef" : '', 
    "Mushrooms" : '',
    "Sausage" : '',
    "1/2 Bacon" : '',
    "1/2 Green Peppers" : '',
    "1/2 Ham" : '', 
    "1/2 Hamburger" : '',
    "1/2 Mushrooms" : '',
    "1/2 Onions" : '',
    "1/2 Pepper Rings" : '', 
    "1/2 Pepperoni" : '', 
    "1/2 Sausage" : '', 
    "Chicken" : '', 
    "Pepper Rings" : '', 
    "Pepperoni" : '', 
    "Pineapple" : ''
}
menu_name_list = list(menu.keys())
menu_value_list = list(menu.values())

wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active


num_rows = sheet_obj.max_row

for i in range(num_rows, 0, -1):
    cell_obj = sheet_obj.cell(row=i, column=2)
    if cell_obj.value not in menu_name_list: 
        sheet_obj.delete_rows(i)
        print(f"Row {i} deleted")
num_rows = sheet_obj.max_row
for i in range(num_rows, 0, -1):
    cell_obj = sheet_obj.cell(row=i, column=8)
    cell_obj.value = menu_value_list[i]
    print(menu_value_list[i+1])
print(menu_name_list)
print(menu_value_list)

wb_obj.save(path)
print("Workbook saved successfully.")
